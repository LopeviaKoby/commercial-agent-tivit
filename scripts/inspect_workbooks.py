from __future__ import annotations

import json
import math
import re
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT_DIR / "data" / "input"
OUTPUT_DIR = ROOT_DIR / "data" / "output"
INVENTORY_PATH = OUTPUT_DIR / "workbook_inventory.json"
QUALITY_PATH = OUTPUT_DIR / "data_quality_report.json"
ROADMAP_PATH = ROOT_DIR / "docs" / "sources" / "Roadmap.md"
QUERIES_DOCX_PATH = ROOT_DIR / "docs" / "sources" / "Consultas habituales de ventas (1).docx"
QUERIES_TXT_PATH = ROOT_DIR / "docs" / "sources" / "Consultas habituales de ventas.txt"

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}
HEADER_SCAN_ROWS = 12
HEADER_KEYWORDS = (
    "acv",
    "budget",
    "status",
    "fase",
    "stage",
    "país",
    "pais",
    "country",
    "data",
    "date",
    "nome",
    "razão",
    "razao",
    "class",
    "lob",
    "bu",
    "partner",
    "parceiro",
    "tvt",
)
IDENTIFIER_KEYWORDS = (
    "id",
    "codigo",
    "código",
    "code",
    "cod",
    "tvt",
)
MONEY_KEYWORDS = (
    "acv",
    "reais",
    "revenue",
    "budget",
    "bud",
    "quota",
    "cotação",
    "cotacao",
    "valor",
    "amount",
    "total",
)
SUMMARY_KEYWORDS = ("summary", "summary", "resumen", "resumo", "top", "pareto", "pivot")
SOURCE_KEYWORDS = ("pipe", "ganad", "actual", "detail", "detalle", "base")


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def safe_str(value: Any) -> str:
    if is_missing(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    return str(value).strip()


def normalize_header(value: Any) -> str:
    text = safe_str(value)
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text)
    return text.casefold()


def header_tokens(value: str) -> set[str]:
    normalized = normalize_header(value)
    if not normalized:
        return set()
    return {token for token in re.split(r"[^0-9A-Za-zÀ-ÿ]+", normalized) if token}


def looks_numeric(value: Any) -> bool:
    return parse_decimal_like(value) is not None


def parse_decimal_like(value: Any) -> float | None:
    if is_missing(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    text = safe_str(value)
    if not text:
        return None

    normalized = text.replace("\u00a0", "").replace(" ", "")
    normalized = normalized.replace("R$", "").replace("$", "")
    normalized = normalized.replace("%", "")
    normalized = normalized.replace("(", "-").replace(")", "")

    if "," in normalized and "." in normalized:
        if normalized.rfind(",") > normalized.rfind("."):
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", "")
    elif "," in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")

    try:
        return float(normalized)
    except ValueError:
        return None


def parse_date_like(value: Any) -> datetime | None:
    if is_missing(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, datetime):
        return value

    text = safe_str(value)
    if not text:
        return None

    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def infer_value_type(values: list[Any]) -> str:
    sample = [value for value in values if not is_missing(value)]
    if not sample:
        return "empty"

    numeric_hits = sum(1 for value in sample if parse_decimal_like(value) is not None)
    date_hits = sum(1 for value in sample if parse_date_like(value) is not None)
    bool_hits = sum(1 for value in sample if isinstance(value, bool))
    total = len(sample)

    if bool_hits == total:
        return "boolean"
    if numeric_hits == total:
        as_numbers = [parse_decimal_like(value) for value in sample]
        if all(number is not None and number.is_integer() for number in as_numbers):
            return "integer"
        return "decimal"
    if date_hits == total:
        return "date"
    if numeric_hits / total >= 0.8:
        return "mostly_decimal"
    if date_hits / total >= 0.8:
        return "mostly_date"
    if numeric_hits > 0 or date_hits > 0:
        return "mixed"
    return "string"


def anonymize_value(value: Any) -> str:
    if is_missing(value):
        return "<empty>"

    if isinstance(value, bool):
        return str(value).lower()

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "<number>"

    parsed_date = parse_date_like(value)
    if parsed_date is not None:
        return parsed_date.date().isoformat()

    text = safe_str(value)
    if text.startswith("="):
        return "<formula>"

    numeric_candidate = parse_decimal_like(text)
    if numeric_candidate is not None:
        return "<number>"

    if "@" in text:
        return "<email_like_text>"

    if re.fullmatch(r"[A-Za-z0-9\-_./]+", text):
        return f"<code_like:{len(text)}>"

    words = [segment for segment in re.split(r"\s+", text) if segment]
    if len(words) >= 2:
        return f"<text_words:{len(words)}>"
    return f"<text:{len(text)}>"


def header_keyword_score(values: list[str]) -> int:
    score = 0
    for value in values:
        lowered = value.casefold()
        if any(keyword in lowered for keyword in HEADER_KEYWORDS):
            score += 2
    return score


def detect_header_row(preview: pd.DataFrame) -> dict[str, Any]:
    best_row = 0
    best_score = float("-inf")
    header_values: list[str] = []

    scan_limit = min(HEADER_SCAN_ROWS, len(preview.index))
    for row_index in range(scan_limit):
        row_values = [safe_str(value) for value in preview.iloc[row_index].tolist()]
        non_empty = [value for value in row_values if value]
        if len(non_empty) < 2:
            score = -100.0 + row_index
        else:
            numeric_hits = sum(1 for value in non_empty if looks_numeric(value))
            unique_ratio = len({value.casefold() for value in non_empty}) / len(non_empty)
            alpha_hits = sum(1 for value in non_empty if re.search(r"[A-Za-zÀ-ÿ]", value))
            score = (
                len(non_empty) * 3
                + unique_ratio * 4
                + alpha_hits * 1.5
                + header_keyword_score(non_empty)
                - numeric_hits * 2.5
                - row_index * 0.4
            )
        if score > best_score:
            best_score = score
            best_row = row_index
            header_values = row_values

    return {
        "row_index_zero_based": best_row,
        "row_number_excel": best_row + 1,
        "score": round(best_score, 2),
        "values": header_values,
    }


def load_preview(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return pd.read_csv(
            path,
            header=None,
            nrows=HEADER_SCAN_ROWS,
            dtype=object,
            encoding="utf-8-sig",
        )

    return pd.read_excel(
        path,
        sheet_name=sheet_name,
        header=None,
        nrows=HEADER_SCAN_ROWS,
        dtype=object,
        engine="openpyxl",
    )


def load_data_frame(path: Path, header_row: int, sheet_name: str | None = None) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        frame = pd.read_csv(
            path,
            header=None,
            skiprows=header_row + 1,
            dtype=object,
            encoding="utf-8-sig",
        )
    else:
        frame = pd.read_excel(
            path,
            sheet_name=sheet_name,
            header=None,
            skiprows=header_row + 1,
            dtype=object,
            engine="openpyxl",
        )

    frame = frame.dropna(axis=1, how="all")
    frame.columns = list(range(frame.shape[1]))
    return frame


def repeated_headers(headers: list[str]) -> list[str]:
    counter = Counter(normalize_header(header) for header in headers if normalize_header(header))
    return sorted(header for header, count in counter.items() if count > 1)


def classify_sheet_role(
    workbook_name: str,
    sheet_name: str,
    row_count: int,
    column_count: int,
    formula_cells_count: int,
) -> str:
    lowered_name = f"{workbook_name} {sheet_name}".casefold()
    if any(keyword in lowered_name for keyword in SUMMARY_KEYWORDS):
        return "summary"
    if any(keyword in lowered_name for keyword in SOURCE_KEYWORDS):
        return "source"
    if formula_cells_count > 0 and row_count <= 200:
        return "summary"
    if row_count >= 20 and column_count >= 5:
        return "source"
    return "auxiliary"


def detect_identifier_candidates(
    column_profiles: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []

    for profile in column_profiles:
        lowered_name = profile["original_name"].casefold()
        tokens = header_tokens(lowered_name)
        unique_ratio = (
            profile["distinct_non_null_values"] / profile["non_null_count"]
            if profile["non_null_count"]
            else 0.0
        )
        has_identifier_name = any(keyword in tokens for keyword in IDENTIFIER_KEYWORDS)
        looks_like_identifier = has_identifier_name or unique_ratio >= 0.95
        if not looks_like_identifier:
            continue

        if profile["duplicate_non_null_values"] <= 0:
            continue

        candidates.append(
            {
                "column_position": profile["column_position"],
                "original_name": profile["original_name"],
                "non_null_count": profile["non_null_count"],
                "distinct_non_null_values": profile["distinct_non_null_values"],
                "duplicate_non_null_values": profile["duplicate_non_null_values"],
                "uniqueness_ratio": round(unique_ratio, 4),
                "reason": "name_based" if has_identifier_name else "uniqueness_based",
            }
        )

    return candidates


def profile_column(series: pd.Series, original_name: str, column_position: int) -> dict[str, Any]:
    values = series.tolist()
    non_null_values = [value for value in values if not is_missing(value)]
    distinct_values = {safe_str(value) for value in non_null_values if safe_str(value)}
    examples = [anonymize_value(value) for value in non_null_values[:50]]
    unique_examples = list(dict.fromkeys(example for example in examples if example != "<empty>"))
    inferred_type = infer_value_type(non_null_values[:300])
    tokens = header_tokens(original_name)
    looks_date = inferred_type in {"date", "mostly_date"} or any(
        token in tokens for token in {"data", "date", "fecha", "closing", "criacao", "criação"}
    )
    looks_monetary = not looks_date and any(keyword in tokens for keyword in MONEY_KEYWORDS)

    return {
        "column_position": column_position,
        "column_letter": get_column_letter(column_position),
        "original_name": original_name,
        "inferred_type": inferred_type,
        "null_percentage": round((len(values) - len(non_null_values)) / len(values) * 100, 2)
        if values
        else 100.0,
        "non_null_count": len(non_null_values),
        "distinct_non_null_values": len(distinct_values),
        "duplicate_non_null_values": max(len(non_null_values) - len(distinct_values), 0),
        "example_values": unique_examples[:3],
        "is_fully_empty": len(non_null_values) == 0,
        "looks_monetary": looks_monetary,
        "looks_date": looks_date,
    }


def exact_duplicate_rows_count(frame: pd.DataFrame) -> int:
    if frame.empty:
        return 0
    normalized = frame.map(safe_str)
    return int(normalized.duplicated().sum())


def inspect_excel_sheet(
    workbook_path: Path,
    workbook_name: str,
    sheet_name: str,
    metadata_workbook: Any,
    values_workbook: Any,
) -> dict[str, Any]:
    preview = load_preview(workbook_path, sheet_name)
    header_info = detect_header_row(preview)
    data_frame = load_data_frame(
        workbook_path,
        header_row=header_info["row_index_zero_based"],
        sheet_name=sheet_name,
    )
    headers = [
        safe_str(value) or f"unnamed_column_{index + 1}"
        for index, value in enumerate(header_info["values"][: data_frame.shape[1]])
    ]

    if len(headers) < data_frame.shape[1]:
        missing_headers = data_frame.shape[1] - len(headers)
        headers.extend(
            f"unnamed_column_{index + len(headers) + 1}" for index in range(missing_headers)
        )

    data_frame.columns = headers[: data_frame.shape[1]]
    data_frame = data_frame.dropna(axis=0, how="all").reset_index(drop=True)

    column_profiles = [
        profile_column(data_frame.iloc[:, index], headers[index], index + 1)
        for index in range(data_frame.shape[1])
    ]
    empty_columns = [
        profile["original_name"] for profile in column_profiles if profile["is_fully_empty"]
    ]

    worksheet = metadata_workbook[sheet_name]
    values_sheet = values_workbook[sheet_name]
    formula_samples: list[dict[str, Any]] = []
    formula_cells_count = 0
    formula_headers: Counter[str] = Counter()
    cached_formula_values = 0

    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells_count += 1
                if cell.row > header_info["row_number_excel"] and cell.column <= len(headers):
                    formula_headers[headers[cell.column - 1]] += 1
                cached_value = values_sheet[cell.coordinate].value
                if cached_value is not None:
                    cached_formula_values += 1
                if len(formula_samples) < 5:
                    header_name = headers[cell.column - 1] if cell.column <= len(headers) else ""
                    formula_samples.append(
                        {
                            "cell": cell.coordinate,
                            "column_letter": get_column_letter(cell.column),
                            "header": header_name,
                            "cached_value_present": cached_value is not None,
                        }
                    )

    table_names = []
    if getattr(worksheet, "tables", None):
        table_names = sorted(worksheet.tables.keys())

    merged_ranges = [str(merged_range) for merged_range in worksheet.merged_cells.ranges]
    sheet_role = classify_sheet_role(
        workbook_name=workbook_name,
        sheet_name=sheet_name,
        row_count=len(data_frame),
        column_count=len(headers),
        formula_cells_count=formula_cells_count,
    )

    return {
        "sheet_name": sheet_name,
        "sheet_state": worksheet.sheet_state,
        "sheet_role": sheet_role,
        "approximate_rows": int(worksheet.max_row),
        "approximate_columns": int(worksheet.max_column),
        "parsed_data_rows": int(len(data_frame)),
        "parsed_columns": int(len(headers)),
        "possible_header_row": header_info["row_number_excel"],
        "header_detection_score": header_info["score"],
        "original_headers": headers,
        "repeated_headers": repeated_headers(headers),
        "fully_empty_columns": empty_columns,
        "columns": column_profiles,
        "null_percentages": {
            profile["original_name"]: profile["null_percentage"] for profile in column_profiles
        },
        "exact_duplicate_rows": exact_duplicate_rows_count(data_frame),
        "possible_repeated_identifiers": detect_identifier_candidates(column_profiles),
        "formulas": {
            "formula_cells_count": formula_cells_count,
            "cached_formula_values_count": cached_formula_values,
            "formula_headers": formula_headers.most_common(),
            "sample_cells": formula_samples,
        },
        "merged_cells": {
            "merged_ranges_count": len(merged_ranges),
            "sample_ranges": merged_ranges[:5],
        },
        "excel_tables": {
            "count": len(table_names),
            "names": table_names,
        },
        "anonymized_examples": {
            profile["original_name"]: profile["example_values"] for profile in column_profiles
        },
    }


def inspect_csv_file(path: Path) -> dict[str, Any]:
    preview = load_preview(path)
    header_info = detect_header_row(preview)
    data_frame = load_data_frame(path, header_row=header_info["row_index_zero_based"])
    headers = [
        safe_str(value) or f"unnamed_column_{index + 1}"
        for index, value in enumerate(header_info["values"][: data_frame.shape[1]])
    ]
    data_frame.columns = headers[: data_frame.shape[1]]
    data_frame = data_frame.dropna(axis=0, how="all").reset_index(drop=True)
    column_profiles = [
        profile_column(data_frame.iloc[:, index], headers[index], index + 1)
        for index in range(data_frame.shape[1])
    ]

    return {
        "file_name": path.name,
        "file_path": str(path.relative_to(ROOT_DIR)),
        "file_size_bytes": path.stat().st_size,
        "file_type": "csv",
        "sheets": [
            {
                "sheet_name": "csv_data",
                "sheet_state": "visible",
                "sheet_role": "source",
                "approximate_rows": int(len(data_frame) + header_info["row_number_excel"]),
                "approximate_columns": int(data_frame.shape[1]),
                "parsed_data_rows": int(len(data_frame)),
                "parsed_columns": int(data_frame.shape[1]),
                "possible_header_row": header_info["row_number_excel"],
                "header_detection_score": header_info["score"],
                "original_headers": headers,
                "repeated_headers": repeated_headers(headers),
                "fully_empty_columns": [
                    profile["original_name"]
                    for profile in column_profiles
                    if profile["is_fully_empty"]
                ],
                "columns": column_profiles,
                "null_percentages": {
                    profile["original_name"]: profile["null_percentage"]
                    for profile in column_profiles
                },
                "exact_duplicate_rows": exact_duplicate_rows_count(data_frame),
                "possible_repeated_identifiers": detect_identifier_candidates(column_profiles),
                "formulas": {
                    "formula_cells_count": 0,
                    "cached_formula_values_count": 0,
                    "formula_headers": [],
                    "sample_cells": [],
                },
                "merged_cells": {"merged_ranges_count": 0, "sample_ranges": []},
                "excel_tables": {"count": 0, "names": []},
                "anonymized_examples": {
                    profile["original_name"]: profile["example_values"]
                    for profile in column_profiles
                },
            }
        ],
    }


def inspect_workbook(path: Path) -> dict[str, Any]:
    if path.suffix.lower() == ".csv":
        return inspect_csv_file(path)

    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    values_workbook = load_workbook(path, read_only=False, data_only=True, keep_links=False)

    try:
        sheets = [
            inspect_excel_sheet(
                workbook_path=path,
                workbook_name=path.name,
                sheet_name=sheet_name,
                metadata_workbook=workbook,
                values_workbook=values_workbook,
            )
            for sheet_name in workbook.sheetnames
        ]
    finally:
        workbook.close()
        values_workbook.close()

    return {
        "file_name": path.name,
        "file_path": str(path.relative_to(ROOT_DIR)),
        "file_size_bytes": path.stat().st_size,
        "file_type": path.suffix.lower().lstrip("."),
        "sheet_names": workbook.sheetnames if "workbook" in locals() else [],
        "visible_sheets": [
            sheet["sheet_name"] for sheet in sheets if sheet["sheet_state"] == "visible"
        ],
        "hidden_sheets": [
            sheet["sheet_name"] for sheet in sheets if sheet["sheet_state"] != "visible"
        ],
        "sheets": sheets,
    }


def load_useful_document_context() -> dict[str, Any]:
    roadmap_text = ROADMAP_PATH.read_text(encoding="utf-8") if ROADMAP_PATH.exists() else ""
    queries_path = QUERIES_DOCX_PATH if QUERIES_DOCX_PATH.exists() else QUERIES_TXT_PATH
    queries_text = queries_path.read_text(encoding="utf-8") if queries_path.exists() else ""

    return {
        "roadmap_available": ROADMAP_PATH.exists(),
        "queries_source_path": str(queries_path.relative_to(ROOT_DIR))
        if queries_path.exists()
        else None,
        "queries_docx_missing": not QUERIES_DOCX_PATH.exists(),
        "roadmap_mentions": sorted(
            {
                keyword
                for keyword in (
                    "Resumen",
                    "2026 Actual",
                    "Pipe Total",
                    "TAXA_Budget",
                    "DE_PARA BU",
                    "Metas por comercial",
                )
                if keyword.casefold() in roadmap_text.casefold()
                or keyword.casefold() in queries_text.casefold()
            }
        ),
    }


def build_quality_report(
    inventory: dict[str, Any], document_context: dict[str, Any]
) -> dict[str, Any]:
    workbooks = inventory["workbooks"]
    sheet_summaries = [
        {
            "workbook": workbook["file_name"],
            "sheet": sheet["sheet_name"],
            "sheet_role": sheet["sheet_role"],
            "parsed_data_rows": sheet["parsed_data_rows"],
            "parsed_columns": sheet["parsed_columns"],
            "exact_duplicate_rows": sheet["exact_duplicate_rows"],
            "possible_repeated_identifiers": sheet["possible_repeated_identifiers"],
            "fully_empty_columns": sheet["fully_empty_columns"],
            "repeated_headers": sheet["repeated_headers"],
            "formula_cells_count": sheet["formulas"]["formula_cells_count"],
            "merged_ranges_count": sheet["merged_cells"]["merged_ranges_count"],
            "table_count": sheet["excel_tables"]["count"],
        }
        for workbook in workbooks
        for sheet in workbook["sheets"]
    ]

    issues: list[dict[str, Any]] = []
    for workbook in workbooks:
        for sheet in workbook["sheets"]:
            if sheet["exact_duplicate_rows"] > 0:
                issues.append(
                    {
                        "severity": "medium",
                        "category": "duplicate_rows",
                        "workbook": workbook["file_name"],
                        "sheet": sheet["sheet_name"],
                        "detail": f"Found {sheet['exact_duplicate_rows']} exact duplicate rows.",
                    }
                )
            if sheet["repeated_headers"]:
                issues.append(
                    {
                        "severity": "medium",
                        "category": "repeated_headers",
                        "workbook": workbook["file_name"],
                        "sheet": sheet["sheet_name"],
                        "detail": (
                            f"Repeated headers detected: {', '.join(sheet['repeated_headers'])}."
                        ),
                    }
                )
            if sheet["fully_empty_columns"]:
                issues.append(
                    {
                        "severity": "low",
                        "category": "empty_columns",
                        "workbook": workbook["file_name"],
                        "sheet": sheet["sheet_name"],
                        "detail": (
                            f"Fully empty columns: {', '.join(sheet['fully_empty_columns'])}."
                        ),
                    }
                )
            for identifier in sheet["possible_repeated_identifiers"]:
                issues.append(
                    {
                        "severity": "medium",
                        "category": "repeated_identifier_candidate",
                        "workbook": workbook["file_name"],
                        "sheet": sheet["sheet_name"],
                        "detail": (
                            f"Possible identifier column '{identifier['original_name']}' repeats "
                            f"{identifier['duplicate_non_null_values']} times."
                        ),
                    }
                )

    return {
        "generated_at": datetime.now(tz=UTC).isoformat(),
        "document_context": document_context,
        "sheet_summaries": sheet_summaries,
        "issues": issues,
    }


def serialize_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    input_files = sorted(
        path
        for path in INPUT_DIR.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
    )

    document_context = load_useful_document_context()
    inventory = {
        "generated_at": datetime.now(tz=UTC).isoformat(),
        "document_context": document_context,
        "workbooks": [inspect_workbook(path) for path in input_files],
    }
    quality_report = build_quality_report(inventory, document_context)

    serialize_json(INVENTORY_PATH, inventory)
    serialize_json(QUALITY_PATH, quality_report)
    print(f"Inventory written to {INVENTORY_PATH}")
    print(f"Quality report written to {QUALITY_PATH}")


if __name__ == "__main__":
    main()
